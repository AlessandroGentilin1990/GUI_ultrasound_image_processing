import json
import os
from dataclasses import dataclass
from typing import Optional, Tuple, List
import numpy as np
from PIL import Image
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
from matplotlib.patches import Rectangle
import matplotlib.lines as mlines
from matplotlib.lines import Line2D
from scipy.ndimage import gaussian_filter1d
import openpyxl
from openpyxl import Workbook
from tkinter import filedialog, messagebox
import scipy.signal


# ---------------------------- Data containers ----------------------------
@dataclass
class ROI:
    x: int
    y: int
    w: int
    h: int

    def as_tuple(self) -> Tuple[int, int, int, int]:
        return (self.x, self.y, self.w, self.h)

    @staticmethod
    def from_tuple(t: Tuple[int, int, int, int]) -> "ROI":
        return ROI(int(t[0]), int(t[1]), int(t[2]), int(t[3]))

@dataclass
class Parameters:
    h_seconds_per_pixel: Optional[float] = None
    v_cm_per_pixel: Optional[float] = None
    sub_roi_1: Optional[ROI] = None
    sub_roi_2: Optional[ROI] = None

    def to_json(self) -> dict:
        return {
            "h_seconds_per_pixel": self.h_seconds_per_pixel,
            "v_cm_per_pixel": self.v_cm_per_pixel,
            "sub_roi_1": self.sub_roi_1.as_tuple() if self.sub_roi_1 else None,
            "sub_roi_2": self.sub_roi_2.as_tuple() if self.sub_roi_2 else None,
        }

    @staticmethod
    def from_json(d: dict) -> "Parameters":
        p = Parameters()
        p.h_seconds_per_pixel = d.get("h_seconds_per_pixel", None)
        p.v_cm_per_pixel = d.get("v_cm_per_pixel", None)
        p.sub_roi_1 = ROI.from_tuple(tuple(d["sub_roi_1"])) if d.get("sub_roi_1") else None
        p.sub_roi_2 = ROI.from_tuple(tuple(d["sub_roi_2"])) if d.get("sub_roi_2") else None
        return p

# ---------------------------- Utility functions ----------------------------
def pil_to_numpy(img: Image.Image) -> np.ndarray:
    if img.mode != "RGB":
        img = img.convert("RGB")
    return np.array(img)

# ---------------------------- Main App ----------------------------
class UltrasoundGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Ultrasound Data Extractor by GentilinLab")

        # State
        self.image_path: Optional[str] = None
        self.img_np_full: Optional[np.ndarray] = None
        self.params = Parameters()

        # Matplotlib figure
        self.fig = Figure(figsize=(8, 6), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.ax.axis('off')
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.root)
        self.canvas_widget = self.canvas.get_tk_widget()
        self.canvas_widget.grid(row=0, column=0, columnspan=1, sticky="nsew")
        self.toolbar = NavigationToolbar2Tk(self.canvas, self.root, pack_toolbar=False)
        self.toolbar.update()
        self.toolbar.grid(row=1, column=0, sticky="ew")

        # Controls
        control_frame = tk.Frame(self.root)
        control_frame.grid(row=0, column=1, rowspan=2, sticky="ns")
        control_frame.columnconfigure(0, weight=1)

        self.btn_import = tk.Button(control_frame, text="Import Image", command=self.import_image)
        self.btn_import.grid(row=0, column=0, padx=6, pady=6, sticky="ew")

        self.btn_calib_h = tk.Button(control_frame, text="Horizontal Calibration (s)", command=lambda: self.start_line_selection("horizontal"), state="disabled")
        self.btn_calib_h.grid(row=1, column=0, padx=6, pady=6, sticky="ew")

        self.btn_calib_v = tk.Button(control_frame, text="Vertical Calibration (cm)", command=lambda: self.start_line_selection("vertical"), state="disabled")
        self.btn_calib_v.grid(row=2, column=0, padx=6, pady=6, sticky="ew")

        self.btn_select_sub_rois = tk.Button(control_frame, text="Select 2 sub-ROIs", command=self.select_two_sub_rois, state="disabled")
        self.btn_select_sub_rois.grid(row=3, column=0, padx=6, pady=6, sticky="ew")

        self.btn_save_params = tk.Button(control_frame, text="Save parameters", command=self.save_params, state="disabled")
        self.btn_save_params.grid(row=4, column=0, padx=6, pady=6, sticky="ew")

        self.btn_load_params = tk.Button(control_frame, text="Import parameters", command=self.load_params)
        self.btn_load_params.grid(row=5, column=0, padx=6, pady=6, sticky="ew")

        self.btn_process = tk.Button(control_frame, text="Process", command=self.process, state="disabled")
        self.btn_process.grid(row=6, column=0, padx=6, pady=6, sticky="ew")

        self.btn_data_analysis = tk.Button(control_frame, text="Data Analysis", command=self.data_analysis)
        self.btn_data_analysis.grid(row=7, column=0, padx=6, pady=6, sticky="ew")


        self.status_var = tk.StringVar(value="Ready.")
        self.status_label = tk.Label(control_frame, textvariable=self.status_var, anchor="w", justify="left")
        self.status_label.grid(row=8, column=0, padx=6, pady=12, sticky="ew")

        self.btn_data_analysis = tk.Button(control_frame, text="Manual Analysis", command=self.manual_analysis)
        self.btn_data_analysis.grid(row=9, column=0, padx=6, pady=6, sticky="ew")


        self.root.rowconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)

        self.pending_points: List[Tuple[float, float]] = []
        self.current_patch: Optional[Rectangle] = None
        self.cid_click = None

        self.horizontal_line: Optional[mlines.Line2D] = None
        self.vertical_line: Optional[mlines.Line2D] = None

    # -------------------- Image display --------------------
    def draw_image(self):
        self.ax.clear()
        self.ax.axis('off')
        if self.img_np_full is not None:
            self.ax.imshow(self.img_np_full)
        # Draw sub-ROIs
        for r in [self.params.sub_roi_1, self.params.sub_roi_2]:
            if r:
                self.ax.add_patch(Rectangle((r.x, r.y), r.w, r.h, fill=False, color="red", linewidth=2))
        self.canvas.draw_idle()

        # ridisegna le linee di calibrazione se presenti
        if self.horizontal_line:
            self.ax.add_line(self.horizontal_line)
        if self.vertical_line:
            self.ax.add_line(self.vertical_line)

    # -------------------- Import --------------------
    def import_image(self):
        path = filedialog.askopenfilename(title="Select image", filetypes=[("Image files","*.png;*.jpg;*.jpeg;*.bmp;*.tif;*.tiff")])
        if not path: return
        try:
            img = Image.open(path)
            self.img_np_full = pil_to_numpy(img)
            self.status_var.set(f"Imported image: {os.path.basename(path)} ({self.img_np_full.shape[1]}x{self.img_np_full.shape[0]})")
            self.enable_controls()
            self.draw_image()
        except Exception as e:
            messagebox.showerror("Error", f"Unable to open image:\n{e}")

    def enable_controls(self):
        for b in [self.btn_calib_h, self.btn_calib_v, self.btn_select_sub_rois, self.btn_save_params]:
            b.config(state="normal")
        self.update_process_button()

    # -------------------- Process --------------------
    def process(self):
        self.process_roi_window()

    # -------------------- Nuova finestra di elaborazione ROI --------------------

#prova con salva excel
    def process_roi_window(self):
        if not (self.params.sub_roi_1 and self.params.sub_roi_2):
            messagebox.showwarning("Attention", "First select the two sub-ROIs.")
            return

        proc_win = tk.Toplevel(self.root)
        proc_win.title("Sub-ROI processing")
        proc_win.state('zoomed')  # massimizza finestra

        # Intercetta la chiusura della finestra per non chiudere la GUI principale
        proc_win.protocol("WM_DELETE_WINDOW", proc_win.destroy)


        # Copia le ROI
        roi_imgs = [
            self.img_np_full[self.params.sub_roi_1.y:self.params.sub_roi_1.y + self.params.sub_roi_1.h,
            self.params.sub_roi_1.x:self.params.sub_roi_1.x + self.params.sub_roi_1.w].copy(),
            self.img_np_full[self.params.sub_roi_2.y:self.params.sub_roi_2.y + self.params.sub_roi_2.h,
            self.params.sub_roi_2.x:self.params.sub_roi_2.x + self.params.sub_roi_2.w].copy()
        ]

        main_frame = tk.Frame(proc_win)
        main_frame.pack(fill='both', expand=True)

        # Slider soglia
        slider = tk.Scale(main_frame, from_=0, to=255, orient='horizontal', label='Theshold')
        slider.pack(side='top', fill='x', padx=10, pady=5)

        # Slider sigma
        sigma_slider = tk.Scale(main_frame, from_=0.5, to=10, resolution=0.1, orient='horizontal', label='Sigma filter')
        sigma_slider.set(3)
        sigma_slider.pack(side='top', fill='x', padx=10, pady=5)

        # Pulsante Extract Data
        btn_extract = tk.Button(main_frame, text="Extract Data", command=lambda: extract_data())
        btn_extract.pack(side='top', padx=10, pady=5)

        # Figure matplotlib
        fig_proc = Figure(figsize=(12, 16), dpi=100)
        axs = [
            fig_proc.add_subplot(4, 1, 1),
            fig_proc.add_subplot(4, 1, 2),
            fig_proc.add_subplot(4, 1, 3),
            fig_proc.add_subplot(4, 1, 4)
        ]
        for ax in axs:
            ax.axis('off')

        canvas = FigureCanvasTkAgg(fig_proc, master=main_frame)
        canvas.get_tk_widget().pack(side='top', fill='both', expand=True)

        # Funzione per calcolare linea di contorno
        def contour_line(bw_img, sigma=3):
            h, w = bw_img.shape
            contour = np.zeros(w)
            for x in range(w):
                col = bw_img[:, x]
                transitions = np.where(np.diff(col.astype(int)) != 0)[0]
                if len(transitions) > 0:
                    contour[x] = transitions.mean()
                else:
                    contour[x] = h // 2
            from scipy.ndimage import gaussian_filter1d
            return gaussian_filter1d(contour, sigma=sigma)

        # Aggiornamento binarizzazione e contorno
        contours_current = [None, None]  # memorizza le linee gialle

        def update_threshold(val=None):
            threshold = int(slider.get())
            sigma = sigma_slider.get()
            for i in range(2):
                ax = axs[i * 2]
                img_np = roi_imgs[i]
                gray = np.dot(img_np[..., :3], [0.299, 0.587, 0.114])
                bw = np.where(gray > threshold, 0, 255).astype(np.uint8)
                ax.clear()
                ax.axis('off')
                ax.imshow(bw, cmap='gray')

                contour = contour_line(bw, sigma=sigma)
                contours_current[i] = contour  # salva per l'export
                ax.plot(np.arange(bw.shape[1]), contour, color='yellow', linewidth=2)

                ax_orig = axs[i * 2 + 1]
                ax_orig.clear()
                ax_orig.axis('off')
                ax_orig.imshow(roi_imgs[i])

            canvas.draw_idle()

        slider.config(command=update_threshold)
        sigma_slider.config(command=update_threshold)
        slider.set(128)
        update_threshold()



        def extract_data():
            if contours_current[0] is None or contours_current[1] is None:
                messagebox.showwarning("Attention", "Refresh the view before extracting data.")
                return

            from openpyxl import Workbook
            wb = Workbook()

            # Definisci i nomi dei fogli
            sheet_names = ["Above", "Below"]

            for i, contour in enumerate(contours_current):
                # Crea il foglio con il nome corretto
                if i == 0:
                    ws = wb.active
                    ws.title = sheet_names[i]
                else:
                    ws = wb.create_sheet(title=sheet_names[i])

                # Inserisci intestazione
                ws.append(["time", "velocity"])

                n = len(contour)
                x_scale = self.params.h_seconds_per_pixel if hasattr(self.params, 'h_seconds_per_pixel') else 1
                y_scale = self.params.v_cm_per_pixel if hasattr(self.params, 'v_cm_per_pixel') else 1

                for col in range(n):
                    t = col * x_scale
                    if i == 0:
                        # Primo ROI: zero in basso, valori positivi verso l’alto
                        y = (roi_imgs[i].shape[0] - contour[col]) * y_scale
                    else:
                        # Secondo ROI: zero in alto, valori negativi verso il basso
                        y = -contour[col] * y_scale
                    ws.append([t, y])

            # Rimuove sheet di default se ancora presente
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if path:
                wb.save(path)
                messagebox.showinfo("Success", f"Data extracted in {path}")

    def update_process_button(self):
        if any([self.params.h_seconds_per_pixel, self.params.v_cm_per_pixel, self.params.sub_roi_1, self.params.sub_roi_2]):
            self.btn_process.config(state="normal")
        else:
            self.btn_process.config(state="disabled")

    # -------------------- Data Analysis --------------------
    def data_analysis(self):
        # Apri finestra pop-up
        da_win = tk.Toplevel(self.root)
        da_win.title("Data Analysis - Import Excel file")
        da_win.geometry("1200x800")  # aumentato lo spazio per area statistiche

        # Frame superiore per pulsante import
        top_frame = tk.Frame(da_win)
        top_frame.pack(side='top', fill='x', pady=5)

        # Frame destro per mostrare statistiche
        right_frame = tk.Frame(da_win)
        right_frame.pack(side='right', fill='y', padx=5, pady=5)
        stats_text = tk.Text(right_frame, width=40)
        stats_text.pack(fill='y', expand=True)

        t_above = v_above = t_below = v_below = None
        filename = None
        excel_path = None  # salva il percorso del file importato

        def import_excel():
            nonlocal t_above, v_above, t_below, v_below, filename, excel_path
            path = filedialog.askopenfilename(
                title="Select Excel file",
                filetypes=[("Excel files", "*.xlsx;*.xls")]
            )
            if not path:
                return

            excel_path = path  # <-- SALVA IL PERCORSO DEL FILE IMPORTATO
            filename = os.path.splitext(os.path.basename(path))[0]

            try:
                wb = openpyxl.load_workbook(path, data_only=True)
                if "Above" not in wb.sheetnames or "Below" not in wb.sheetnames:
                    messagebox.showerror("Error", "The Excel file must contain the 'Above' and 'Below' sheets.")
                    return

                above_sheet = wb["Above"]
                below_sheet = wb["Below"]

                def sheet_to_arrays(sheet):
                    times, velocities = [], []
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[0] is not None and row[1] is not None:
                            times.append(row[0])
                            velocities.append(row[1])
                    return np.array(times), np.array(velocities)

                t_above, v_above = sheet_to_arrays(above_sheet)
                t_below, v_below = sheet_to_arrays(below_sheet)

                # Pulisce figure esistenti
                for ax in axs:
                    ax.clear()

                # Grafico Above
                axs[0].plot(t_above, v_above, color='blue')
                axs[0].set_title("Above")
                axs[0].set_xlabel("Time [s]")
                axs[0].set_ylabel("Velocity [cm/s]")

                # Grafico Below
                axs[1].plot(t_below, v_below, color='red')
                axs[1].set_title("Below")
                axs[1].set_xlabel("Time [s]")
                axs[1].set_ylabel("Velocity [cm/s]")

                fig_da.tight_layout()
                canvas.draw_idle()

            except Exception as e:
                messagebox.showerror("Error", f"Unable to read Excel file:\n{e}")

        def calculate_statistics():
            nonlocal t_above, v_above, t_below, v_below, filename, excel_path
            if v_above is None or v_below is None or excel_path is None:
                messagebox.showerror("Error", "Please import a valid Excel file first.")
                return

            # Indici principali da mostrare nella GUI (10)
            def compute_main_indices(signal):
                indices = {
                    "Mean": np.mean(signal),
                    "Median": np.median(signal),
                    "Std Dev": np.std(signal),
                    "Variance": np.var(signal),
                    "Min": np.min(signal),
                    "Max": np.max(signal),
                    "Range": np.ptp(signal),
                    "Skewness": scipy.stats.skew(signal),
                    "Kurtosis": scipy.stats.kurtosis(signal),
                    "RMS": np.sqrt(np.mean(signal ** 2))
                }
                return indices

            # Tutti e 50 indici da salvare nell'Excel
            def compute_all_indices(signal, times):
                from scipy.signal import find_peaks
                indices = {}
                n = len(signal)
                indices["Mean"] = np.mean(signal)
                indices["Median"] = np.median(signal)
                indices["Std Dev"] = np.std(signal)
                indices["Variance"] = np.var(signal)
                indices["Min"] = np.min(signal)
                indices["Max"] = np.max(signal)
                indices["Range"] = np.ptp(signal)
                indices["Skewness"] = scipy.stats.skew(signal)
                indices["Kurtosis"] = scipy.stats.kurtosis(signal)
                indices["RMS"] = np.sqrt(np.mean(signal ** 2))
                indices["IQR"] = np.percentile(signal, 75) - np.percentile(signal, 25)
                indices["Sum"] = np.sum(signal)
                indices["Integral"] = np.trapz(signal, times)
                indices["AUC"] = np.trapz(np.abs(signal), times)
                indices["Median Absolute Deviation"] = np.median(np.abs(signal - np.median(signal)))
                indices["Max Absolute"] = np.max(np.abs(signal))
                indices["Min Absolute"] = np.min(np.abs(signal))
                indices["Peak Count"] = len(find_peaks(signal)[0])
                indices["Peak Mean"] = np.mean(signal[find_peaks(signal)[0]]) if len(find_peaks(signal)[0]) > 0 else 0
                indices["Peak Std"] = np.std(signal[find_peaks(signal)[0]]) if len(find_peaks(signal)[0]) > 0 else 0
                indices["Energy"] = np.sum(signal ** 2)
                indices["Mean Abs Deviation"] = np.mean(np.abs(signal - np.mean(signal)))
                indices["Median Abs Deviation"] = np.median(np.abs(signal - np.median(signal)))
                indices["Max Slope"] = np.max(np.diff(signal))
                indices["Min Slope"] = np.min(np.diff(signal))
                indices["Mean Slope"] = np.mean(np.diff(signal))
                indices["Std Slope"] = np.std(np.diff(signal))
                indices["Variance Slope"] = np.var(np.diff(signal))
                indices["RMS Slope"] = np.sqrt(np.mean(np.diff(signal) ** 2))
                indices["Skewness Slope"] = scipy.stats.skew(np.diff(signal))
                indices["Kurtosis Slope"] = scipy.stats.kurtosis(np.diff(signal))
                indices["Sum Positive"] = np.sum(signal[signal > 0])
                indices["Sum Negative"] = np.sum(signal[signal < 0])
                indices["Count Positive"] = np.sum(signal > 0)
                indices["Count Negative"] = np.sum(signal < 0)
                indices["Max Positive"] = np.max(signal[signal > 0]) if np.any(signal > 0) else 0
                indices["Min Negative"] = np.min(signal[signal < 0]) if np.any(signal < 0) else 0
                indices["Mean Positive"] = np.mean(signal[signal > 0]) if np.any(signal > 0) else 0
                indices["Mean Negative"] = np.mean(signal[signal < 0]) if np.any(signal < 0) else 0
                indices["Std Positive"] = np.std(signal[signal > 0]) if np.any(signal > 0) else 0
                indices["Std Negative"] = np.std(signal[signal < 0]) if np.any(signal < 0) else 0
                indices["Variance Positive"] = np.var(signal[signal > 0]) if np.any(signal > 0) else 0
                indices["Variance Negative"] = np.var(signal[signal < 0]) if np.any(signal < 0) else 0
                indices["RMS Positive"] = np.sqrt(np.mean(signal[signal > 0] ** 2)) if np.any(signal > 0) else 0
                indices["RMS Negative"] = np.sqrt(np.mean(signal[signal < 0] ** 2)) if np.any(signal < 0) else 0
                indices["Skewness Positive"] = scipy.stats.skew(signal[signal > 0]) if np.any(signal > 0) else 0
                indices["Skewness Negative"] = scipy.stats.skew(signal[signal < 0]) if np.any(signal < 0) else 0
                indices["Kurtosis Positive"] = scipy.stats.kurtosis(signal[signal > 0]) if np.any(signal > 0) else 0
                indices["Kurtosis Negative"] = scipy.stats.kurtosis(signal[signal < 0]) if np.any(signal < 0) else 0
                return indices

            # 10 indici principali per la GUI
            indices_above_gui = compute_main_indices(v_above)
            indices_below_gui = compute_main_indices(v_below)

            stats_text.delete("1.0", tk.END)
            stats_text.insert(tk.END, "Above:\n")
            for key, value in indices_above_gui.items():
                stats_text.insert(tk.END, f"{key}: {value:.4f}\n")
            stats_text.insert(tk.END, "\nBelow:\n")
            for key, value in indices_below_gui.items():
                stats_text.insert(tk.END, f"{key}: {value:.4f}\n")

            # 50 indici per Excel
            indices_above_excel = compute_all_indices(v_above, t_above)
            indices_below_excel = compute_all_indices(v_below, t_below)

            wb_stats = openpyxl.Workbook()
            wb_stats.remove(wb_stats.active)

            def write_sheet(name, indices_dict):
                ws = wb_stats.create_sheet(name)
                ws.append(["Index", "Value"])
                for k, v in indices_dict.items():
                    ws.append([k, v])

            write_sheet("Above", indices_above_excel)
            write_sheet("Below", indices_below_excel)

            save_path = os.path.join(os.path.dirname(excel_path), f"{filename}_stats.xlsx")
            wb_stats.save(save_path)
            messagebox.showinfo("Success", f"Statistics saved in {save_path}")

        btn_import_excel = tk.Button(top_frame, text="Import Excel file", command=import_excel)
        btn_import_excel.pack(side='left', padx=10)

        btn_calc_stats = tk.Button(top_frame, text="Calculate Statistics", command=calculate_statistics)
        btn_calc_stats.pack(side='left', padx=10)

        # Frame principale per figure
        fig_da = Figure(figsize=(10, 6), dpi=100)
        axs = [
            fig_da.add_subplot(2, 1, 1),
            fig_da.add_subplot(2, 1, 2)
        ]

        canvas = FigureCanvasTkAgg(fig_da, master=da_win)
        canvas.get_tk_widget().pack(side='top', fill='both', expand=True)


    # -------------------- Calibration --------------------
    def start_line_selection(self, orientation: str):
        if self.img_np_full is None:
            return
        # stop eventuali eventi precedenti
        if self.cid_click:
            self.canvas.mpl_disconnect(self.cid_click)
            self.cid_click = None
        self.pending_points = []
        if self.current_patch:
            try:
                self.current_patch.remove()
            except Exception:
                pass
            self.current_patch = None

        # connetti nuovo evento
        self.cid_click = self.canvas.mpl_connect(
            "button_press_event",
            lambda e: self.on_line_click(e, orientation)
        )
        self.status_var.set(f"Select two points for calibration {orientation}.")

    def on_line_click(self, event, orientation):
        if event.inaxes != self.ax:
            return
        if event.xdata is None or event.ydata is None:
            return  # click fuori dal canvas

        self.pending_points.append((event.xdata, event.ydata))

        if len(self.pending_points) == 1:
            # marker primo punto
            if self.current_patch:
                self.current_patch.remove()
            self.current_patch = Line2D([event.xdata], [event.ydata],
                                        marker="o", color="blue", linestyle="None")
            self.ax.add_line(self.current_patch)
            self.canvas.draw_idle()

        elif len(self.pending_points) == 2:
            # rimuovi marker e disegna linea finale
            if self.current_patch:
                self.current_patch.remove()
            (x1, y1), (x2, y2) = self.pending_points
            if x1 is None or y1 is None or x2 is None or y2 is None:
                messagebox.showerror("Error", "Invalid coordinates")
                self.pending_points = []
                return

            dist = np.hypot(x2 - x1, y2 - y1)
            if dist <= 0:
                messagebox.showerror("Error", "Zero distance")
                self.pending_points = []
                self.current_patch = None
                return

            # disegna linea finale
            line = Line2D([x1, x2], [y1, y2], color="blue", linewidth=2)
            self.ax.add_line(line)
            self.canvas.draw_idle()

            # salva la linea come attributo in base all'orientamento
            if orientation == "horizontal":
                self.horizontal_line = line
            else:
                self.vertical_line = line

            # chiedi valore reale
            if orientation == "horizontal":
                val = simpledialog.askfloat("Horizontal Calibration",
                                            "Duration (s):", minvalue=1e-12)
                if val is not None:
                    self.params.h_seconds_per_pixel = val / dist
                    self.status_var.set(
                        f"Horizontal Calibration: {self.params.h_seconds_per_pixel:.6g} s/pixel")
            else:
                val = simpledialog.askfloat("Vertical Calibration",
                                            "Height (cm):", minvalue=1e-12)
                if val is not None:
                    self.params.v_cm_per_pixel = val / dist
                    self.status_var.set(
                        f"Vertical Calibration: {self.params.v_cm_per_pixel:.6g} cm/pixel")

            # pulizia
            self.pending_points = []
            if self.cid_click:
                self.canvas.mpl_disconnect(self.cid_click)
            self.cid_click = None

    def cleanup_line(self):
        if self.cid_click: self.canvas.mpl_disconnect(self.cid_click)
        self.cid_click=None
        self.pending_points=[]



    def select_two_sub_rois(self):
        if self.img_np_full is None: return
        self.params.sub_roi_1=None
        self.params.sub_roi_2=None
        self.pending_points=[]
        self.cid_click = self.canvas.mpl_connect("button_press_event", self.on_sub_roi_click)
        self.status_var.set("Select two points for the 1st sub-ROI")

        self.current_sub_roi = 1

    def on_sub_roi_click(self, event):
        if event.inaxes != self.ax: return
        self.pending_points.append((event.xdata, event.ydata))
        if len(self.pending_points)==2:
            x0, y0 = self.pending_points[0]
            x1, y1 = self.pending_points[1]
            roi = ROI(min(x0,x1), min(y0,y1), abs(x1-x0), abs(y1-y0))
            if self.current_sub_roi==1:
                self.params.sub_roi_1=roi
                self.status_var.set("1st sub-ROI selected. Select the 2nd sub-ROI")
                self.pending_points=[]
                self.current_sub_roi=2
            else:
                self.params.sub_roi_2=roi
                self.status_var.set("2nd sub-ROI selected")
                self.cleanup_sub_roi()
                self.update_process_button()
            self.draw_image()

    def cleanup_sub_roi(self):
        if self.cid_click: self.canvas.mpl_disconnect(self.cid_click)
        self.cid_click=None
        self.pending_points=[]
        self.current_sub_roi=None

    # -------------------- Save / Load --------------------
    def save_params(self):
        if not any([self.params.h_seconds_per_pixel,self.params.v_cm_per_pixel,self.params.sub_roi_1,self.params.sub_roi_2]):
            messagebox.showwarning("Nothing to save","Define at least one parameter first")
            return
        d=self.params.to_json()
        d["_image_path"]=self.image_path
        path=filedialog.asksaveasfilename(title="Save parameters", defaultextension=".json", filetypes=[("JSON","*.json")])
        if not path: return
        try:
            with open(path,"w") as f: json.dump(d,f,indent=2)
            self.status_var.set(f"Saved parameters: {os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("Error",f"Unable to save:\n{e}")

    def load_params(self):
        path=filedialog.askopenfilename(title="Import parameters", filetypes=[("JSON","*.json")])
        if not path: return
        try:
            with open(path,"r") as f: d=json.load(f)
            self.params=Parameters.from_json(d)
            if self.img_np_full is None and d.get("_image_path") and os.path.isfile(d["_image_path"]):
                img=Image.open(d["_image_path"])
                self.img_np_full=pil_to_numpy(img)
                self.image_path=d["_image_path"]
                self.enable_controls()
            self.draw_image()
            self.status_var.set(f"Imported parameters: {os.path.basename(path)}")
            self.update_process_button()
        except Exception as e:
            messagebox.showerror("Error",f"Unable to import:\n{e}")

    # -------------------------- Manual --------------
    # -------------------- Manual Analysis --------------------

    def manual_analysis(self):
        import numpy as np
        from scipy.interpolate import CubicSpline
        from openpyxl import Workbook

        if self.img_np_full is None:
            messagebox.showwarning("Attention", "Import an image first.")
            return
        if not (self.params.sub_roi_1 and self.params.sub_roi_2):
            messagebox.showwarning("Attention", "Define the two sub-ROIs first.")
            return

        # Estrai le due ROI
        roi_imgs = [
            self.img_np_full[self.params.sub_roi_1.y:self.params.sub_roi_1.y + self.params.sub_roi_1.h,
            self.params.sub_roi_1.x:self.params.sub_roi_1.x + self.params.sub_roi_1.w].copy(),
            self.img_np_full[self.params.sub_roi_2.y:self.params.sub_roi_2.y + self.params.sub_roi_2.h,
            self.params.sub_roi_2.x:self.params.sub_roi_2.x + self.params.sub_roi_2.w].copy()
        ]

        win = tk.Toplevel(self.root)
        win.title("Manual Analysis - Draw Contours")
        win.geometry("800x1000")

        win.protocol("WM_DELETE_WINDOW", win.destroy)

        # Frame per pulsanti in alto
        btn_frame = tk.Frame(win)
        btn_frame.pack(side='top', fill='x', pady=5)

        lines = [[], []]  # punti disegnati per ciascun ROI
        current_roi = [0]  # ROI attivo
        prev_point = [None]  # punto precedente per linee continue

        # Creazione del canvas con le due ROI
        fig = Figure(figsize=(8, 10), dpi=100)
        axs = [fig.add_subplot(2, 1, 1), fig.add_subplot(2, 1, 2)]
        for i, ax in enumerate(axs):
            ax.axis('off')
            ax.imshow(roi_imgs[i])
            ax.set_title(f"ROI {i + 1}")

        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.get_tk_widget().pack(fill='both', expand=True)

        draw_active = [False]  # stato disegno: True=può disegnare, False=non può

        def start_draw():
            draw_active[0] = not draw_active[0]  # alterna stato
            if draw_active[0]:
                self.status_var.set("Draw the yellow lines with the mouse. Press 'Draw' to disable them.")
                canvas.mpl_connect("motion_notify_event", draw_line)
                canvas.mpl_connect("button_press_event", start_line)
                canvas.mpl_connect("button_release_event", end_line)
            else:
                self.status_var.set("Drawing is disabled. Press 'Draw' to reactivate.")



        def start_line(event):
            if event.inaxes is None or event.button != 1:
                return
            prev_point[0] = (event.xdata, event.ydata)
            roi_idx = current_roi[0]
            lines[roi_idx].append([prev_point[0]])  # nuova linea come lista di punti

        def draw_line(event):
            if event.inaxes is None or prev_point[0] is None:
                return
            roi_idx = current_roi[0]
            x0, y0 = prev_point[0]
            x1, y1 = event.xdata, event.ydata
            axs[roi_idx].plot([x0, x1], [y0, y1], color='yellow')
            lines[roi_idx][-1].append((x1, y1))
            prev_point[0] = (x1, y1)
            canvas.draw_idle()

        def end_line(event):
            prev_point[0] = None


        def next_roi():
            # alterna tra ROI 0 e ROI 1
            current_roi[0] = 1 - current_roi[0]
            self.status_var.set(f"Draw the outline on the ROI {current_roi[0] + 1}")

        def save_lines():
            if not any(lines):
                messagebox.showwarning("Attention", "There are no lines to save.")
                return

            from openpyxl import Workbook
            from scipy.interpolate import CubicSpline
            import numpy as np

            wb = Workbook()

            # Array per memorizzare i dati interpolati ad alta risoluzione
            high_res_data = [{}, {}]  # Uno per ogni ROI

            for i, roi_lines in enumerate(lines):
                # Foglio originale con punti disegnati
                ws_original = wb.create_sheet(f"ROI_{i + 1}")
                ws_original.append(["x", "y", "y_spline"])

                # Dizionario x -> y unico per ROI
                x_points = {}
                for line in roi_lines:
                    for x, y in line:
                        x_val = x * self.params.h_seconds_per_pixel if self.params.h_seconds_per_pixel else x

                        # gestione Y invertita a seconda del ROI
                        if i == 0:  # ROI superiore
                            y_val = (roi_imgs[0].shape[
                                         0] - y) * self.params.v_cm_per_pixel if self.params.v_cm_per_pixel else (
                                        roi_imgs[0].shape[0] - y)
                        else:  # ROI inferiore
                            y_val = -y * self.params.v_cm_per_pixel if self.params.v_cm_per_pixel else -y

                        x_points[x_val] = y_val

                # Ordina gli x
                xs_sorted = np.array(sorted(x_points.keys()))
                ys_sorted = np.array([x_points[x] for x in xs_sorted])

                # Spline cubica
                if len(xs_sorted) >= 4:
                    cs = CubicSpline(xs_sorted, ys_sorted, extrapolate=True)
                    ys_spline = cs(xs_sorted)
                else:
                    ys_spline = ys_sorted

                # Scrivi nel foglio originale
                for x_val, y_val, y_spl in zip(xs_sorted, ys_sorted, ys_spline):
                    ws_original.append([x_val, y_val, y_spl])

                # Calcola l'intervallo completo di X per alta risoluzione
                if xs_sorted.size > 0:
                    min_x = xs_sorted.min()
                    max_x = xs_sorted.max()

                    # Crea array X ad alta risoluzione (un punto per ogni pixel temporale)
                    # Calcola il numero di punti basato sulla risoluzione temporale
                    if self.params.h_seconds_per_pixel:
                        # Calcola il range temporale totale
                        total_time = max_x - min_x
                        # Numero di punti = tempo totale / risoluzione temporale
                        num_points = int(total_time / self.params.h_seconds_per_pixel)
                    else:
                        # Se non c'è calibrazione, usa 1000 punti
                        num_points = 1000

                    x_high_res = np.linspace(min_x, max_x, num_points)

                    # Calcola Y ad alta risoluzione con spline
                    if len(xs_sorted) >= 4:
                        y_high_res = cs(x_high_res)
                    else:
                        # Interpolazione lineare se non abbastanza punti per spline
                        y_high_res = np.interp(x_high_res, xs_sorted, ys_sorted)

                    # Salva per i fogli Above/Below
                    high_res_data[i] = {'x': x_high_res, 'y': y_high_res}

            # Crea fogli Above e Below con dati ad alta risoluzione
            sheet_names = ["Above", "Below"]
            for i in range(2):
                if high_res_data[i]:  # Se ci sono dati
                    ws_high_res = wb.create_sheet(sheet_names[i])
                    ws_high_res.append(["time", "velocity"])

                    for t, v in zip(high_res_data[i]['x'], high_res_data[i]['y']):
                        ws_high_res.append([t, v])

            # Rimuovi sheet di default
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if path:
                wb.save(path)
                messagebox.showinfo("Success", f"Lines saved in {path}")



        # Pulsanti
        tk.Button(btn_frame, text="Draw", command=start_draw).pack(side='left', padx=5)
        btn_next = tk.Button(btn_frame, text="Next ROI", command=next_roi)
        btn_next.pack(side='left', padx=5)
        tk.Button(btn_frame, text="Save", command=save_lines).pack(side='left', padx=5)


# ---------------------------- Main ----------------------------
def main():
    root=tk.Tk()
    app=UltrasoundGUI(root)
    root.mainloop()

if __name__=="__main__":
    main()

